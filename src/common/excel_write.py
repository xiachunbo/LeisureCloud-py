#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

openpyxl_data = [
    ('1','ACT_EVT_LOG', '事件日志表','LOG_NR_'),
    ('2','ACT_EVT_LOG', '事件日志表','TYPE_'),
    ('3','ACT_EVT_LOG', '事件日志表','PROC_DEF_ID_'),
    ('4','ACT_EVT_LOG', '事件日志表','PROC_INST_ID_'),
    ('5','ACT_EVT_LOG', '事件日志表','EXECUTION_ID_'),
    ('6','ACT_EVT_LOG', '事件日志表','TASK_ID_'),
    ('7','ACT_EVT_LOG', '事件日志表','TIME_STAMP_'),
    ('8','ACT_EVT_LOG', '事件日志表','USER_ID_' ),
    ('9','ACT_EVT_LOG', '事件日志表','DATA_'),
    ('10','ACT_EVT_LOG', '事件日志表','LOCK_OWNER_'),
    ('11','ACT_EVT_LOG', '事件日志表','LOCK_TIME_'),
    ('12','ACT_EVT_LOG', '事件日志表','IS_PROCESSED_')
]
output_file_name = 'test.xlsx'


def save_excel(target_list, output_file_name):
    """
    将数据写入xlsx文件
    """
    if not output_file_name.endswith('.xlsx'):
        output_file_name += '.xlsx'

    # 创建一个workbook对象，而且会在workbook中至少创建一个表worksheet
    #wb = openpyxl.Workbook()
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    wb = openpyxl.load_workbook(r'/Users/xiaobobo/Desktop/test.xlsx')
    wb1 = wb.create_sheet(title='Mysheet1', index=3000)
    #title_data = ('a', 'b', 'c', 'd', 'e', 'f')
    #target_list.insert(0, title_data)
    rows = len(target_list)
    lines = len(target_list[0])
    for i in range(rows):
        for j in range(lines):
            wb1.cell(row=i+3 + 1, column=j + 1).value = target_list[i][j]

    # 保存表格
    wb.save(filename=output_file_name)
save_excel(openpyxl_data, output_file_name)