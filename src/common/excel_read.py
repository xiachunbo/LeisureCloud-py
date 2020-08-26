#!/usr/bin/env python
# -*- coding: utf-8 -*-
from aetypes import end

import openpyxl

input_file_name = '/Users/xiaobobo/Desktop/test.xlsx'


def read_excel(input_file_name):
    """
    从xlsx文件中读取数据
    """
    workbook = openpyxl.load_workbook(input_file_name)
    print(workbook)
    # 可以使用workbook对象的sheetnames属性获取到excel文件中哪些表有数据
    print(workbook.sheetnames)
    workbook.active = 0  # 设置active参数，即工作表索引值
    table = workbook.active
    print(table)
    rows = table.max_row
    cols = table.max_column

    for row in range(rows):
        for col in range(cols):
            data = table.cell(row + 1, col + 1).value
            print(data)


read_excel(input_file_name)